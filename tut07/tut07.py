from datetime import datetime
start_time = datetime.now()

def modification(file):
    workbook = load_workbook(filename='output/'+file)
    sheet = workbook.active
    #removing the column heading we given previously
    for coloumns in ['AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ']:
        sheet[coloumns + '1'] = ''
    side=Side(border_style='thin', color="000000")
    border=Border(top=side,bottom=side,left=side,right=side)
    #giving border for ranking
    for cell in sheet['N1':'AF' + str(ranges+2)]:
        for c in cell:
            c.border = border
    #giving border for count of Rank mod values
    for cell in sheet['AC' + str(ranges+4):'AE' + str(ranges+12)]:
        for c in cell:
            c.border = border
    #giving border for overall transition count
    for cell in sheet['AI4':'AQ12']:
        for c in cell:
            c.border = border
    #giving border for mod transition count
    for i in range(ranges):
        for cell in sheet['AI' + str(18 + 14*i):'AQ' + str(26 + 14*i)]:
            for c in cell:
                c.border = border
    #giving border for longest subsequence length
    for cell in sheet['AS1':'AU9']:
        for c in cell:
            c.border = border
    #giving border for longest subsequence range
    for cell in sheet['AW1':'AY' + str(17 + sum([sheet['AU' + str(i)].value for i in range(2, 10)]))]:
        for c in cell:
            c.border = border
    
    #giving colour for rank1 ocant
    style = DifferentialStyle(fill=PatternFill(bgColor='FFFF00'))
    rank_rule = Rule(type="expression", dxf=style)
    rank_rule.formula = ['W2=1']
    sheet.conditional_formatting.add("W2:AE" + str(ranges + 2), rank_rule)
    #giving colour for highest transition count
    max_rule = Rule(type="expression", dxf=style)
    max_rule.formula = ['AJ4=MAX($AJ4:$AQ4)']
    sheet.conditional_formatting.add("AJ4:AQ12", max_rule)
    for i in range(ranges ):
        rule = Rule(type="expression", dxf=style)
        rule.formula = ['AJ' + str(19 + 14*i) + '=MAX($AJ' + str(19 + 14*i) + ':$AQ' + str(19 + 14*i) + ')']
        sheet.conditional_formatting.add("AJ" + str(19 + 14*i) + ":AQ" + str(26 + 14*i), rule)  
    workbook.save('output/'+file)
#Help https://youtu.be/N6PBd4XdnEw
def octant_range_names(mod=5000):
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

    #initialised a list for getting range values
    range_list=[0]
    #initialised x as variable for getting number of mod ranges
    x=int(len(df)/mod)
    global ranges 
    ranges= x+1
    #adding range values in the list
    for i in range (x):
        range_list.append(mod*(i+1))
    range_list.append(len(df))
    for i in range(len(range_list)-1):
        for octant in ['1','-1','2','-2','3','-3','4','-4']:
            #counting the counts of octants in the mod ranges
            df.at[1+i,octant]=df['Octant'].iloc[range_list[i]:range_list[i+1]].value_counts()[int(octant)]
        df.at[1+i,'Octant ID']=str(range_list[i])+" - "+ str(range_list[i+1]-1)  
    
    #initialised 2 lists for taking the overall counts of the octant values
    rank_list_overall_1=[]
    rank_list_overall_2=[]    
    #The value of overall octant count is added to both the lists initialised above
    for y in ['1','-1','2','-2','3','-3','4','-4']:
        rank_list_overall_1.append(df[y][0])
        rank_list_overall_2.append(df[y][0])
    #sorting one of the list(rank_list_overall_2) while the other one remains unsorted
    rank_list_overall_2.sort(reverse=True)
    #Initialised a list called final_list_1 for ranking the octants
    final_list_1=[]
    for i in range(0,len(rank_list_overall_2)):
        #rank_list_overall_1 is unsorted list while rank_list_overall_2 is sorted list
        #what the below line do's is: comparing the value is sorted list to that of unsorted list and
        # returns of the index of that value in unsorted array
        #The index is then appended to the final_list_1 list
        final_list_1.append((rank_list_overall_2.index(rank_list_overall_1[i]))+1) 
    # for i in range(1,9):
    #     #adding the rank values in respective columns
    #     df.at[0,'Rank'+str(i)]=final_list_1[i-1]
    q=1
    for oct in ['1','-1','2','-2','3','-3','4','-4']:
        df.at[0,'Rank Octant '+oct]=final_list_1[q-1]
        #adding the index of octant having rank 1
        if final_list_1[q-1]==1:
            z=q-1
        for k,x in enumerate([1,-1,2,-2,3,-3,4,-4]):
            try:
                #when the k value becomes equal to that of the z value assigned, octant value corresponding to that is added to variable named value 
                if k==z:
                    value=x
            except:
                continue
        q+=1
    #Adding the rank 1 Octant ID and Octant name
    df.at[0,'Rank1 OctantID']=value
    df.at[0,'Rank1 Octant Name']=octant_name_id_mapping.get(str(value))

    #The above steps for overall octant count is repeated for ranking the octant counts for different mod ranges
    #adding the octant counts in two lists, sorting one list, comparing the values and adding the index values to one list
    #the index values is then printed
    for i in range(len(range_list)-1):
        mod_list_1=[]
        mod_list_2=[]
        for y in ['1','-1','2','-2','3','-3','4','-4']:
            mod_list_1.append(df[y][1+i])
            mod_list_2.append(df[y][1+i])
        mod_list_2.sort(reverse=True)
        final_list_2=[]
        for z in range(0,len(mod_list_2)):
            final_list_2.append((mod_list_2.index(mod_list_1[z]))+1)        
        w=1
        for oct in ['1','-1','2','-2','3','-3','4','-4']:
            df.at[1+i,'Rank Octant '+oct]=final_list_2[w-1]    
            if final_list_2[w-1]==1:
                a=w-1
            for k,x in enumerate([1,-1,2,-2,3,-3,4,-4]):
                try: 
                    if k==a:
                        df.at[1+i,'Rank1 OctantID']=x
                        df.at[1+i,'Rank1 Octant Name']=octant_name_id_mapping.get(str(x))
                except:
                    continue
            w+=1            
    df.at[1+len(range_list),'Rank Octant 4']='Octant ID'
    df.at[1+len(range_list),'Rank Octant -4']='Octant Name'
    df.at[1+len(range_list),'Rank1 OctantID']='Count of Rank 1 Mod Values'

    #initialsed a list containing 8 elements with every elements 8
    count_list=[0,0,0,0,0,0,0,0]
    #Traversing Rank1 OctantID column
    #when the value in the columns becomes equal to 1 the first element in list in increased by one
    #when it becomes equal to -1 second elemnt is increased by 1 and respectively for others
    for i in range(1,len(df['Rank1 OctantID'])):
        if df['Rank1 OctantID'][i]==1:
            count_list[0]+=1
        elif df['Rank1 OctantID'][i]==-1:
              count_list[1]+=1  
        elif df['Rank1 OctantID'][i]==2:
              count_list[2]+=1  
        elif df['Rank1 OctantID'][i]==-2:
              count_list[3]+=1      
        elif df['Rank1 OctantID'][i]==3:
              count_list[4]+=1      
        elif df['Rank1 OctantID'][i]==-3:
              count_list[5]+=1  
        elif df['Rank1 OctantID'][i]==4:
              count_list[6]+=1  
        elif df['Rank1 OctantID'][i]==-4:
              count_list[7]+=1  
    #printing the count of rank 1 mod values
    for x,octval in enumerate([1,-1,2,-2,3,-3,4,-4]):
        df.at[2+len(range_list)+x,'Rank Octant 4']=octval
        df.at[2+len(range_list)+x,'Rank Octant -4']=octant_name_id_mapping.get(str(octval))
        df.at[2+len(range_list)+x,'Rank1 OctantID']=count_list[x]
    df['  ']=''
    range_list[-1]-=1
    df['AH']=''
    df['AI']=''
    df['AJ']=''
    df['AK']=''
    df['AL']=''
    df['AM']=''
    df['AN']=''
    df['AO']=''
    df['AP']=''
    df['AQ']=''

    # # #heading of coloumns
    Row_num=0
    df.at[Row_num,'AI']="Overall Transition Count"
    df.at[Row_num+1,'AJ']="To"
    df.at[Row_num+2,'AI']="Octant"
    df.at[Row_num+3,'AH']="From"
    columns=['AJ','AK','AL','AM','AN','AO','AP','AQ']
    #Octant values printing in row and column manner
    for i,x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
        df.at[Row_num+3+i,'AI']=x
        df.at[Row_num+2,columns[i]]=x
    
    # # #dictionary for storing octant transition values
    Trans_count_dict = {
    1: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0},
    -1: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0},
    2: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0},
    -2: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0},
    3: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0},
    -3: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0},
    4: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0},
    -4: {1: 0,-1: 0,2: 0,-2: 0,3: 0,-3: 0,4: 0,-4: 0}
    }


    for num in range(len(range_list)-1):
        mod_row = Row_num+14+num*(14)
        mod_transitions_count =copy.deepcopy(Trans_count_dict)
        df.at[mod_row,'AI'] = 'Mod Transition Count'
        df.at[mod_row + 1, 'AI'] = str(range_list[num]) + '-' + str(range_list[num+1]-1)
        df.at[mod_row +1, 'AJ']="To"
        df.at[mod_row +2, 'AI'] = "Octant"
        for i,x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
            df.at[mod_row+3+i,'AI']=x
            df.at[mod_row+2,columns[i]]=x
        df.at[mod_row + 3, 'AH'] = "From"
        for i in range(range_list[num], range_list[num+1]):
            mod_transitions_count[df['Octant'][i+1]][df['Octant'][i]]+=1
    #mod transition count
        for x,i in enumerate(mod_transitions_count):
            for idx, j in enumerate(mod_transitions_count[i]):
                df.at[mod_row+3+idx,columns[x]]=mod_transitions_count[i][j] 
    #overall transition count
    for i in range(1, len(df['U'])):
        Trans_count_dict[df['Octant'][i]][df['Octant'][i-1]]+=1
    for x,i in enumerate(Trans_count_dict):
        for idx, j in enumerate(Trans_count_dict[i]):
            df.at[Row_num+3+idx,columns[x]]=Trans_count_dict[i][j]

    df['    ']=''
    #printing octant values
    for i,x in enumerate([1,-1,2,-2,3,-3,4,-4]):
        df.at[i,'octant']=x
    
    #finding longest subsequence and count of longest subsequence 
    for j,octval in enumerate([1,-1,2,-2,3,-3,4,-4]):
        count=0
        maxi=0
        Subsequence_count=0

        #when the octant value in coloumn of Octant becomes equal 
        #to that of octant value obtained through enumeration the count is increased
        for i in range(len(df['Octant'])):
            if df['Octant'][i]==octval:
                count+=1

                #Maximum value is updated when the count becomes greater than current maximum value.
                maxi=max(count,maxi)
            else:
                count=0
        count=0

        #finding count of longest subsequence
        for i in range(len(df['Octant'])):
            if df['Octant'][i]==octval:
                count+=1

                #when the count becomes equal to maximum value obtained previously, subsequent count is increased
                if(count==maxi):
                    Subsequence_count+=1
            else:
                count=0
        
        # Printing the values in coloumns of excel
        try:
            df.at[j,'Longest Subsequence Length']=maxi
            df.at[j,'Count']=Subsequence_count
        except:
            print("An error occured while printing Longest subsequence / Count ")
            break
    df['      ']=''
     #printing octant values
    for i,x in enumerate([1,-1,2,-2,3,-3,4,-4]):
        df.at[i,'octant']=x
    
    #initialized variables for getting row number
    c1=0
    l1=0
    c2=0
    #finding longest subsequence and count of longest subsequence 
    for j,octval in enumerate([1,-1,2,-2,3,-3,4,-4]):
        count=0
        maxi=0
        Subsequence_count=0

        try:
            #when the octant value in coloumn of Octant becomes equal 
            #to that of octant value obtained through enumeration the count is increased
            for i in range(len(df['Octant'])):
                if df['Octant'][i]==octval:
                    count+=1

                    #Maximum value is updated when the count becomes greater than current maximum value.
                    maxi=max(count,maxi)
                else:
                    count=0
            count=0
        except:
            print("Error while finding longest subsequence of octant:",octval)
        #finding count of longest subsequence
        x=0 
        list=[]
        try:
            for i in range(len(df['Octant'])):
                if df['Octant'][i]==octval:
                    count+=1

                    #when the count becomes equal to maximum value obtained previously, subsequent count is increased
                    if(count==maxi):     
                        Subsequence_count+=1

                        # when count becomes equal to longest subsequence value, the index is added to list
                        list.append(i)
                else:
                    count=0
        except:
            print("Error while finding count of longest subsequence of octant :",octval)
        try:
            # Printing the values in coloumns of excel
            df.at[j,'Longest Subsequence Length']=maxi
            df.at[j,'Count']=Subsequence_count
        except:
            print("Error while printing Longest subsequence and count")
        
        try:
            # printing octant values in Octant_2 column and string Time
            df.at[c1,'Octant_2']=octval
            c1+=1
            df.at[c1,'Octant_2']='Time'

            # increasing c1 by total number of subsequence count to match the given format 
            for i in range(Subsequence_count+1):
                c1+=1
            
            # printing longest subsequence value in the column
            df.at[l1,'Longest Subsequence Length 2']=maxi
            l1+=1
            df.at[l1,'Longest Subsequence Length 2']='From'
            l1+=1

            # printing the From time range for all the longest subsequence using the help of list of longest subsequence created before
            for i in range(len(list)):
                df.at[l1,'Longest Subsequence Length 2']=df['T'][list[i]+1-maxi]
                l1+=1
            df.at[c2,'Count_2']=Subsequence_count
            c2+=1
            df.at[c2,'Count_2']='To'
            c2+=1

            # printing the To time for all the longest subsequence using the help of list of longest subsequence created before
            for i in range(len(list)):
                df.at[c2,'Count_2']=df['T'][list[i]]
                c2+=1
        except:
            print("Error while printing time range of longest subsequence")



from platform import python_version
import copy
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule 
ver = python_version()
try:
    import pandas as pd

    for filename in os.listdir('input'):
        f=os.path.join('input',filename)
        if not os.path.isfile(f) or not f.endswith('.xlsx'):
            continue
        df=pd.read_excel(f)
        #computing the average values of U,V and W and adding them to column
        U_Avg  = df['U'].mean()
        df.at[0,'U Avg']=U_Avg
        V_Avg  =df['V'].mean()
        df.at[0,'V Avg']=V_Avg
        W_Avg  = df['W'].mean()
        df.at[0,'W Avg']=W_Avg
        
        #computing U',V' and W'
        df["U'=U - U avg"]=df['U']-df['U Avg'][0]
        df["V'=V - V avg"]=df['V']-df['V Avg'][0]
        df["W'=W - W avg"]=df['W']-df['W Avg'][0]
        #Finding the Octant and adding those in Octant column
        for i in range(len(df)):
            if df["U'=U - U avg"][i]>0 and df["V'=V - V avg"][i]>0 and df["W'=W - W avg"][i]>0:
                df.at[i,"Octant"]=1
            elif df["U'=U - U avg"][i]>0 and df["V'=V - V avg"][i]>0 and df["W'=W - W avg"][i]<0:
                df.at[i,"Octant"]=-1
            elif df["U'=U - U avg"][i]<0 and df["V'=V - V avg"][i]>0 and df["W'=W - W avg"][i]>0:
                df.at[i,"Octant"]=2
            elif df["U'=U - U avg"][i]<0 and df["V'=V - V avg"][i]>0 and df["W'=W - W avg"][i]<0:
                df.at[i,"Octant"]=-2
            elif df["U'=U - U avg"][i]<0 and df["V'=V - V avg"][i]<0 and df["W'=W - W avg"][i]>0:
                df.at[i,"Octant"]=3
            elif df["U'=U - U avg"][i]<0 and df["V'=V - V avg"][i]<0 and df["W'=W - W avg"][i]<0:
                df.at[i,"Octant"]=-3
            elif df["U'=U - U avg"][i]>0 and df["V'=V - V avg"][i]<0 and df["W'=W - W avg"][i]>0:
                df.at[i,"Octant"]=4
            elif df["U'=U - U avg"][i]>0 and df["V'=V - V avg"][i]<0 and df["W'=W - W avg"][i]<0:
                df.at[i,"Octant"]=-4
        mod=5000

        df['']=''
        df.at[0,' ']='Mod '+str(mod)
        df.at[0,'Octant ID']="Overall Count"
        df.reset_index(drop=True) 
        count1,counti,count2,countii,count3,countiii,count4,countiv=0,0,0,0,0,0,0,0

        #Calculating the count of octants (Overall count)
        #Variable count1 takes count of octant 1, Counti takes count of octant -1 and similarly for others.
        # Traversing through the column octant and add one to the respective count variable 
        for i in range(len(df)):
            if df.at[i,'Octant'] == 1:
                count1+=1
            elif df.at[i,'Octant']==-1:
                counti+=1
            elif df.at[i,'Octant']==2:
                count2+=1
            elif df.at[i,'Octant']==-2:
                countii+=1
            elif df.at[i,'Octant']==3:
                count3+=1
            elif df.at[i,'Octant']==-3:
                countiii+=1
            elif df.at[i,'Octant']==4:
                count4+=1
            elif df.at[i,'Octant']==-4:
                countiv+=1

        #Printing the overal counts
        df.at[0,'1']=count1
        df.at[0,'-1']=counti
        df.at[0,'2']=count2
        df.at[0,'-2']=countii
        df.at[0,'3']=count3
        df.at[0,'-3']=countiii

        df.at[0,'4']=count4
        df.at[0,'-4']=countiv
        octant_range_names(mod)
        df.to_excel('output/'+ filename, index=False)  
        modification(filename)

except ImportError:
    print("Module Pandas not Found!")

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))